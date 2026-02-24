"""
CleanLog - 金融数据整合与一致性校验
模拟鼎甲金融客户数据一致性校验场景
"""
import io
import json
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import streamlit as st

# 使用共享组件（modules/ui_components.py，若项目中有 utils 则改为 from utils.ui_components import ...）
from utils.ui_components import render_page_header, render_download_section

DINGJIA_SCENARIO = "模拟鼎甲金融数据整合与一致性校验场景"

# 标准 schema，用于跨平台统一
STANDARD_SCHEMA = [
    "timestamp",
    "amount",
    "currency",
    "category",
    "account",
    "counterparty",
    "transaction_id",
    "raw_source",
]

# 各平台 CSV 列名 → 标准列 映射模板
PLATFORM_SCHEMAS = {
    "alipay": {
        "timestamp": ["创建时间", "交易创建时间", "付款时间"],
        "amount": ["金额"],
        "direction": ["收/支", "收入/支出"],
        "category": ["类型", "交易类型"],
        "counterparty": ["交易对方", "对方账户"],
        "transaction_id": ["订单号", "交易订单号"],
    },
    "wechat": {
        "timestamp": ["交易时间"],
        "amount": ["金额(元)", "金额"],
        "direction": ["收/支"],
        "category": ["交易类型"],
        "counterparty": ["交易对方", "商品"],
        "transaction_id": ["交易单号", "商户单号"],
    },
    "bank": {
        "timestamp": ["交易时间", "交易日期", "记账时间", "交易日期时间"],
        "amount": ["金额", "交易金额", "收入金额", "支出金额"],
        "direction": ["收付标志", "借贷标志"],
        "category": ["摘要", "交易摘要", "交易类型"],
        "counterparty": ["对方户名", "交易对手", "对方账号", "对方名称"],
        "transaction_id": ["流水号", "交易流水号", "参考号"],
    },
}


def _similarity(a: str, b: str) -> float:
    """计算两个字符串的相似度 [0, 1]"""
    if pd.isna(a) or pd.isna(b):
        return 0.0
    sa, sb = str(a).strip(), str(b).strip()
    if not sa or not sb:
        return 1.0 if sa == sb else 0.0
    return SequenceMatcher(None, sa, sb).ratio()


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """确保输出包含标准 schema 所有列"""
    for col in STANDARD_SCHEMA:
        if col not in df.columns:
            df[col] = pd.NA
    return df[STANDARD_SCHEMA].copy()


class CrossPlatformReconciler:
    """
    跨平台对账器 - 模拟鼎甲金融客户数据一致性校验
    支持支付宝/微信/银行 CSV 自适应识别、模糊匹配、单边账检测
    """

    def __init__(
        self,
        amount_tolerance: float = 0.01,
        time_window_minutes: int = 5,
        name_similarity_threshold: float = 0.6,
    ):
        self.amount_tolerance = amount_tolerance
        self.time_window_minutes = time_window_minutes
        self.name_similarity_threshold = name_similarity_threshold

    def _detect_platform(self, df: pd.DataFrame, filename: str = "") -> str:
        """根据列名和文件名识别数据源"""
        cols = set(df.columns)
        fname = filename.lower()

        if "支付宝" in filename or "alipay" in fname:
            return "alipay"
        if "微信" in filename or "wechat" in fname or "wx" in fname:
            return "wechat"
        if any(k in fname for k in ["bank", "银行", "bankcard", "流水"]):
            return "bank"

        # 列名匹配
        for platform, schema in PLATFORM_SCHEMAS.items():
            ts_cands = schema.get("timestamp", [])
            amt_cands = schema.get("amount", []) + schema.get("direction", [])
            if any(c in cols for c in ts_cands) and any(c in cols for c in amt_cands):
                return platform

        return "bank"  # 默认尝试银行格式

    def _build_mapping(self, platform: str, df: pd.DataFrame) -> dict:
        """构建原始列 → 标准列映射"""
        schema = PLATFORM_SCHEMAS.get(platform, PLATFORM_SCHEMAS["bank"])
        mapping = {}

        for std_col, cands in schema.items():
            if std_col == "direction":
                continue
            for c in cands:
                if c in df.columns:
                    mapping[c] = std_col
                    break

        return mapping

    def _parse_amount(self, series: pd.Series, direction_series: Optional[pd.Series] = None) -> pd.Series:
        """解析金额：去逗号、符号，根据收/支设正负"""
        s = series.astype(str).str.replace(",", "", regex=False).str.replace("¥", "", regex=False)
        vals = pd.to_numeric(s, errors="coerce").fillna(0)

        if direction_series is not None:
            direction = direction_series.astype(str)
            vals = vals.where(
                direction.str.contains("收入|收到|收款|贷|收", na=False, regex=True),
                -vals.abs(),
            )
        return vals

    def extract_with_auto_schema(
        self, file_path_or_bytes, filename: str = ""
    ) -> Tuple[pd.DataFrame, str, dict]:
        """
        自适应提取：自动识别平台并映射 schema
        支持 CSV 与 Excel (.xlsx, .xls)
        返回 (DataFrame, platform, column_mapping)
        """
        def _get_readable():
            """返回每次可重新读取的源"""
            if isinstance(file_path_or_bytes, bytes):
                return io.BytesIO(file_path_or_bytes)
            if hasattr(file_path_or_bytes, "getvalue"):
                return io.BytesIO(file_path_or_bytes.getvalue())
            return file_path_or_bytes

        fn_lower = (filename or "").lower()
        is_excel = fn_lower.endswith(".xlsx") or fn_lower.endswith(".xls")

        if is_excel:
            src = _get_readable()
            engine = "openpyxl" if fn_lower.endswith(".xlsx") else None
            df = pd.read_excel(src, engine=engine)
        else:
            for enc in ("utf-8", "gbk", "gb18030"):
                try:
                    src = _get_readable()
                    df = pd.read_csv(src, encoding=enc)
                    break
                except (UnicodeDecodeError, Exception):
                    continue
            else:
                raise ValueError("无法解码文件编码")

        if df.empty:
            return _ensure_columns(df), "unknown", {}

        platform = self._detect_platform(df, filename)
        mapping = self._build_mapping(platform, df)

        rename_map = {k: v for k, v in mapping.items() if k in df.columns}
        df = df.rename(columns=rename_map)

        # 平台标识
        platform_labels = {"alipay": "Alipay", "wechat": "WeChat", "bank": "Bank"}
        df["account"] = platform_labels.get(platform, platform)
        df["raw_source"] = platform
        df["currency"] = "CNY"

        # 时间解析
        if "timestamp" not in df.columns:
            raise ValueError(f"未找到时间列，当前列: {list(df.columns)}")
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
        df = df.dropna(subset=["timestamp"])

        # 金额解析
        dir_col = next(
            (c for c in PLATFORM_SCHEMAS.get(platform, {}).get("direction", []) if c in df.columns),
            None,
        )
        if "amount" in df.columns:
            dir_series = df[dir_col] if dir_col else None
            df["amount"] = self._parse_amount(df["amount"], dir_series)
        elif platform == "bank":
            # 部分银行分开 收入/支出 列
            inc_col = next((c for c in ["收入金额", "贷方金额"] if c in df.columns), None)
            exp_col = next((c for c in ["支出金额", "借方金额"] if c in df.columns), None)
            if inc_col and exp_col:
                inc = pd.to_numeric(df[inc_col].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
                exp = pd.to_numeric(df[exp_col].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
                df["amount"] = inc - exp
            elif inc_col:
                df["amount"] = pd.to_numeric(df[inc_col].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
            elif exp_col:
                df["amount"] = -pd.to_numeric(df[exp_col].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
            else:
                df["amount"] = 0.0

        # 删除临时列
        df = df.drop(columns=[c for c in df.columns if c.startswith("_")], errors="ignore")

        return _ensure_columns(df), platform, {v: k for k, v in rename_map.items()}

    def fuzzy_match(
        self, df: pd.DataFrame
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        模糊匹配算法：金额相同 + 时间±5分钟 + 对方名称相似度
        返回 (matched, suspected_duplicate, unilateral)
        """
        if df.empty or len(df) < 2:
            unilateral = df.copy()
            unilateral["match_status"] = "unilateral"
            return (
                pd.DataFrame(columns=df.columns.tolist() + ["match_status"]),
                pd.DataFrame(columns=df.columns.tolist() + ["match_status"]),
                unilateral,
            )

        df = df.copy()
        df["_idx"] = range(len(df))

        matched_indices = set()
        suspected_indices = set()

        for i in range(len(df)):
            if i in matched_indices or i in suspected_indices:
                continue
            row_i = df.iloc[i]
            ts_i = row_i["timestamp"]
            amt_i = row_i["amount"]
            name_i = str(row_i.get("counterparty", "") or "")

            for j in range(i + 1, len(df)):
                if j in matched_indices or j in suspected_indices:
                    continue
                row_j = df.iloc[j]
                ts_j = row_j["timestamp"]
                amt_j = row_j["amount"]
                name_j = str(row_j.get("counterparty", "") or "")

                # 金额容差
                if abs(amt_i - amt_j) > self.amount_tolerance:
                    continue
                # 时间窗口 ±5 分钟
                if abs((ts_i - ts_j).total_seconds()) > self.time_window_minutes * 60:
                    continue
                # 名称相似度
                sim = _similarity(name_i, name_j)
                if sim < self.name_similarity_threshold:
                    continue

                if sim >= 0.9:
                    matched_indices.add(i)
                    matched_indices.add(j)
                else:
                    suspected_indices.add(i)
                    suspected_indices.add(j)
                break  # 每个 i 只匹配一个 j

        def tag_df(indices: set, status: str) -> pd.DataFrame:
            sub = df[df["_idx"].isin(indices)].copy()
            sub["match_status"] = status
            sub = sub.drop(columns=["_idx"], errors="ignore")
            return sub

        matched = tag_df(matched_indices, "matched")
        suspected = tag_df(suspected_indices, "suspected_duplicate")
        unilateral_indices = set(df["_idx"]) - matched_indices - suspected_indices
        unilateral = tag_df(unilateral_indices, "unilateral")

        return matched, suspected, unilateral

    def consistency_check(
        self, df: pd.DataFrame
    ) -> dict:
        """
        数据一致性校验：金额平衡检查、单边账检测
        """
        result = {
            "amount_balanced": True,
            "total_amount": 0.0,
            "unilateral_count": 0,
            "suspected_duplicate_count": 0,
            "matched_count": 0,
            "warnings": [],
        }

        if df.empty:
            return result

        result["total_amount"] = float(df["amount"].sum())

        if "match_status" not in df.columns:
            matched, suspected, unilateral = self.fuzzy_match(df)
            result["matched_count"] = len(matched)
            result["suspected_duplicate_count"] = len(suspected)
            result["unilateral_count"] = len(unilateral)
        else:
            result["matched_count"] = int((df["match_status"] == "matched").sum())
            result["suspected_duplicate_count"] = int((df["match_status"] == "suspected_duplicate").sum())
            result["unilateral_count"] = int((df["match_status"] == "unilateral").sum())

        if abs(result["total_amount"]) > 100:
            result["amount_balanced"] = False
            result["warnings"].append(f"金额不平衡: 净额 ¥{result['total_amount']:,.2f}，请核对单边账")

        if result["unilateral_count"] > len(df) * 0.5:
            result["warnings"].append("单边账占比过高，建议检查数据完整性")

        return result

    def build_excel_report(
        self,
        matched: pd.DataFrame,
        suspected: pd.DataFrame,
        unilateral: pd.DataFrame,
        consistency: dict,
    ) -> bytes:
        """生成含公式的 Excel 差异报告"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # 无 openpyxl 时退化为基础 Excel
            for engine in ("xlsxwriter", "openpyxl"):
                try:
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine=engine) as w:
                        matched.to_excel(w, sheet_name="匹配成功", index=False)
                        suspected.to_excel(w, sheet_name="疑似重复", index=False)
                        unilateral.to_excel(w, sheet_name="单边账", index=False)
                    return buf.getvalue()
                except ImportError:
                    continue
            return b""

        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for name, data in [
            ("匹配成功", matched),
            ("疑似重复", suspected),
            ("单边账", unilateral),
        ]:
            if name in wb.sheetnames:
                ws = wb[name]
            else:
                ws = wb.create_sheet(name)
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

        # 汇总 sheet（含公式）
        ws_sum = wb.create_sheet("汇总", 0)
        ws_sum["A1"] = "对账汇总"
        ws_sum["A1"].font = Font(bold=True, size=14)
        ws_sum["A2"] = "匹配成功笔数"
        ws_sum["B2"] = len(matched)
        ws_sum["A3"] = "疑似重复笔数"
        ws_sum["B3"] = len(suspected)
        ws_sum["A4"] = "单边账笔数"
        ws_sum["B4"] = len(unilateral)
        ws_sum["A5"] = "总笔数（公式）"
        ws_sum["B5"] = "=B2+B3+B4"
        ws_sum["A6"] = "金额是否平衡"
        ws_sum["B6"] = "是" if consistency.get("amount_balanced", True) else "否"
        ws_sum["A7"] = "净额"
        ws_sum["B7"] = consistency.get("total_amount", 0)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def run_finance_etl(file_paths: list) -> dict:
    """对外接口：基于路径执行财务 ETL（兼容 app.py 原有调用）"""
    reconciler = CrossPlatformReconciler()
    all_dfs = []
    all_mappings = []

    for path in file_paths:
        p = Path(path)
        if not p.exists():
            continue
        try:
            df, platform, mapping = reconciler.extract_with_auto_schema(str(p), p.name)
            if not df.empty:
                all_dfs.append(df)
                all_mappings.append({"file": p.name, "platform": platform, "mapping": mapping})
        except Exception:
            pass

    if not all_dfs:
        return {
            "module": "finance_etl",
            "dingjia_scenario": DINGJIA_SCENARIO,
            "problem_discovery": {"提示": "未成功加载任何账单文件"},
            "cleaning_actions": [],
            "effect_verification": {},
            "details": {},
            "raw_df": pd.DataFrame(),
            "report_json": json.dumps({"error": "no_data"}, ensure_ascii=False),
            "use_streamlit_ui": True,
        }

    merged = pd.concat(all_dfs, ignore_index=True).drop_duplicates()
    matched, suspected, unilateral = reconciler.fuzzy_match(merged)
    consistency = reconciler.consistency_check(
        pd.concat([matched, suspected, unilateral], ignore_index=True)
    )

    excel_bytes = reconciler.build_excel_report(matched, suspected, unilateral, consistency)

    return {
        "module": "finance_etl",
        "dingjia_scenario": DINGJIA_SCENARIO,
        "problem_discovery": {
            "总交易数": f"{len(merged):,}",
            "匹配成功": f"{len(matched)}",
            "疑似重复": f"{len(suspected)}",
            "单边账": f"{len(unilateral)}",
        },
        "cleaning_actions": ["已执行：Schema 自动映射、模糊匹配、一致性校验"],
        "effect_verification": {
            "金额平衡": "是" if consistency["amount_balanced"] else "否",
            "净额 (¥)": f"{consistency['total_amount']:,.2f}",
        },
        "details": {
            "consistency": consistency,
            "mappings": all_mappings,
        },
        "raw_df": merged,
        "matched_df": matched,
        "suspected_df": suspected,
        "unilateral_df": unilateral,
        "column_mappings": all_mappings,
        "excel_report_bytes": excel_bytes,
        "report_json": json.dumps(
            {
                "matched": len(matched),
                "suspected_duplicate": len(suspected),
                "unilateral": len(unilateral),
                "consistency": consistency,
            },
            ensure_ascii=False,
            indent=2,
        ),
        "use_streamlit_ui": True,
    }


def render_finance_etl_ui():
    """Streamlit 界面：多文件上传、映射展示、三栏对账结果、Excel 下载"""
    render_page_header("财务对账", DINGJIA_SCENARIO)

    uploaded_files = st.file_uploader(
        "上传账单文件（支持支付宝/微信/银行 CSV、Excel，可多选）",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=True,
        help="同时上传多个平台账单进行跨平台对账",
    )

    if not uploaded_files:
        st.info("👆 请上传一个或多个 CSV 或 Excel 账单文件")
        return

    if st.button("▶ 开始对账", type="primary"):
        reconciler = CrossPlatformReconciler()
        all_dfs = []
        all_mappings = []

        with st.spinner("正在加载并映射数据..."):
            for f in uploaded_files:
                try:
                    df, platform, mapping = reconciler.extract_with_auto_schema(
                        f.getvalue(), f.name
                    )
                    if not df.empty:
                        all_dfs.append(df)
                        all_mappings.append({
                            "file": f.name,
                            "platform": platform,
                            "mapping": mapping,
                        })
                except Exception as e:
                    st.warning(f"跳过 {f.name}: {e}")

        if not all_dfs:
            st.error("未成功加载任何账单")
            return

        merged = pd.concat(all_dfs, ignore_index=True).drop_duplicates()
        matched, suspected, unilateral = reconciler.fuzzy_match(merged)
        consistency = reconciler.consistency_check(
            pd.concat([matched, suspected, unilateral], ignore_index=True)
        )
        excel_bytes = reconciler.build_excel_report(
            matched, suspected, unilateral, consistency
        )

        # 自动字段映射展示（表格）
        st.subheader("📋 自动字段映射")
        mapping_rows = []
        for m in all_mappings:
            for std_col, orig_col in m["mapping"].items():
                mapping_rows.append({
                    "文件": m["file"],
                    "平台": m["platform"],
                    "标准列": std_col,
                    "原始列": orig_col,
                })
        if mapping_rows:
            st.dataframe(pd.DataFrame(mapping_rows), use_container_width=True, hide_index=True)
        else:
            st.caption("无额外映射（已使用默认列名）")

        # 对账结果三栏
        st.subheader("📊 对账结果")
        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("✅ 匹配成功", len(matched))
            if not matched.empty:
                with st.expander("查看数据", expanded=False):
                    st.dataframe(matched, use_container_width=True, hide_index=True)

        with col2:
            st.metric("⚠️ 疑似重复", len(suspected))
            if not suspected.empty:
                with st.expander("查看数据", expanded=False):
                    st.dataframe(suspected, use_container_width=True, hide_index=True)

        with col3:
            st.metric("❌ 单边账", len(unilateral))
            if not unilateral.empty:
                with st.expander("查看数据", expanded=False):
                    st.dataframe(unilateral, use_container_width=True, hide_index=True)

        # 一致性校验摘要
        if consistency.get("warnings"):
            st.warning("; ".join(consistency["warnings"]))

        # 差异报告下载（Excel）
        st.subheader("📥 下载差异报告")
        st.download_button(
            "下载 Excel 报告（含汇总与公式）",
            excel_bytes,
            file_name="reconciliation_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # 使用共享组件渲染额外下载（JSON/CSV）
        render_download_section(
            {
                "report_json": json.dumps(
                    {
                        "matched": len(matched),
                        "suspected_duplicate": len(suspected),
                        "unilateral": len(unilateral),
                        "consistency": consistency,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                "raw_df": merged,
            },
            "finance_reconciliation",
        )
